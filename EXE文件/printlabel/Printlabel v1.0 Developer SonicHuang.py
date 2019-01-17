import openpyxl as opx
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import sys
import os
import easygui as eg
from easygui import EgStore

while 1:
    readpath = eg.fileopenbox('Where\'s your source?', 'PrintLabel 1.0v --Writer:SonicHuang')
    if readpath == None:
        sys.exit()
    if os.path.splitext(readpath)[1] != '.xlsx':
        eg.msgbox('CAUTION!:\nThe file is not Excel file or the Excel version is too old, the file extension must be xlsx.\nPlease select file again.', 'PrintLabel 1.0v --Writer:SonicHuang')
        continue
    else:
        break
    


class Settings(EgStore):
    def __init__(self, filename):
        self.dimensions = ['28.3', '28.3', '28.3', '12.34', '12.34']
        self.fontstyle = ['宋体', '16', 'y', 'y','00BEBEBE', '宋体', '16', 'y', 'y', '00BEBEBE']
        self.filename = filename
        self.restore()


settingsFilename = os.path.join(os.curdir, 'printlabelsettings.txt' )
settings = Settings(settingsFilename)
dimensions = settings.dimensions
fontstyle = settings.fontstyle
msg = '<<...................Dimensions....................>>\nCheck the Excel Columnr/Row Setting Options for refrence.\nYou should input number from 8 to up for width, 14 to up for height'
title = 'PrintLabel 1.0v --Writer:SonicHuang'
fields = ['1st Row Height', '2nd Row Height', '3rd Row Height', 'A Column Width', 'B Column Width']
msg1 = '<<...................Font....................>>\nCheck the Excel Font Name and Size for your refrence\nColumn Color, check the Hex Color Code, example: Grey color is #BEBEBE, change # to 00, input 00BEBEBE'
fields1 = ['A Column Font Name', 'A Column Font Size','A Conlumn Bold(y/n)', 'A Conlumn Italic(y/n)', 'A Column Color', 'B Column Font Name', 'B Column Font Size', 'B Column Bold(y/n)', 'B Conlumn Italic(y/n)', 'B Column Color']

dimensions = eg.multenterbox(msg, title, fields, dimensions)
if dimensions == None:
    sys.exit()
fontstyle = eg.multenterbox(msg1, title, fields1, fontstyle)
if fontstyle == None:
    sys.exit()



settings.dimensions = dimensions
settings.fontstyle = fontstyle
settings.store()
dict1 = dict(zip(fields, dimensions))
dict2 = dict(zip(fields1, fontstyle))
h1 = dict1['1st Row Height']
h2 = dict1['2nd Row Height']
h3 = dict1['3rd Row Height']
aw = dict1['A Column Width']
bw = dict1['B Column Width']
af = dict2['A Column Font Name']
bf = dict2['B Column Font Name']
afs = dict2['A Column Font Size']
bfs = dict2['B Column Font Size']
acc = dict2['A Column Color']
bcc = dict2['B Column Color']
acb = dict2['A Conlumn Bold(y/n)']
if acb == 'y':
    acb = True
else:
    acb = None
bcb = dict2['B Column Bold(y/n)']
if bcb == 'y':
    bcb = True
else:
    bcb = None
aci = dict2['A Conlumn Italic(y/n)']
if aci == 'y':
    aci = True
else:
    aci = None

bci = dict2['B Conlumn Italic(y/n)']
if bci == 'y':
    bci = True
else:
    bci = None


wbr = opx.load_workbook(readpath)
sheetr = wbr.active
rownum = sheetr[1]

temppath = os.path.join(os.curdir, 'printlabeltempfile')
os.mkdir(temppath)

for r in range(2, sheetr.max_row + 1):
    
    writepath = os.path.join(temppath, 'temp' + str(r) +'.xlsx')
    wbw = opx.Workbook()
    sheetw = wbw.active
    sheetw.row_dimensions[1].height = float(h1)
    sheetw.row_dimensions[2].height = float(h2)
    sheetw.row_dimensions[3].height = float(h3)
    sheetw.column_dimensions['A'].width = aw
    sheetw.column_dimensions['B'].width = bw
    try:
        i = 0
        for num in rownum:
            i += 1
            sheetw['A' + str(i)].font = Font(name = af, size = float(afs), color = acc, bold = acb, italic = aci)
            sheetw['A' + str(i)].alignment = Alignment(horizontal = 'center', vertical = 'center')
            sheetw['A' + str(i)] = num.value
            if i == 3:
                raise Exception('nextround')
    except Exception:
        
        n = 0
        for data in sheetr[r]:
            n += 1
            sheetw['B' + str(n)].font = Font(name = bf, size = float(bfs), color = bcc, bold = bcb, italic = bci)
            sheetw['B' + str(n)].alignment = Alignment(horizontal = 'center', vertical = 'center')
            sheetw['B' + str(n)] = data.value
            if n == 3:
                wbw.save(writepath)
                os.startfile(writepath, 'print')
                continue

askdel = eg.ccbox('System will delet the tempfiles, press any buttom after all the file\'r printed', 'PrintLabel 1.0v --Writer:SonicHuang', ('All files\'r printed', 'Del the tempfiles'))
if askdel == True or askdel == False or askdel == None:
    receive = os.system('rmdir /s /q printlabeltempfile')
    pass



