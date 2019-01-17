import openpyxl as opx
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import sys
import os
import easygui as eg
from easygui import EgStore

def resource_path(relative_path): #这段函数是为了把这个程序用pyintaller打包成exe。参数是相对路径，即在程序所在文件夹的路径。最后在2.0中用这个的方法，同样用ccbox控制时间，这样不会因为临时文件存在而产生错误。每次会产生不同名字的临时文件夹，即使中途有问题临时文件没有被删除，第二次启动程序的时候也不会有问题。而且也不用调用sys来启动cmd黑框。
    if getattr(sys, 'frozen', False): #这是python的内置函数，用于判断第一个参数（对象），是否含有名为第二个参数的属性，如果没有返回第三个参数的值。根据资料，如果sys里面有'frozen'属性，说明运行的是被打包成exe的文件。如果正常运行，py文件，就没有'frozen'属性，则返回False，返回打包前的文件地址。
        base_path = sys._MEIPASS #系统会在运行exe文件时在临时文件夹中新建一个_MEIxxxx，xxxx是随机数字的文件夹，并把所有需要的文件资源存放在里面。当关闭运行exe后，系统自动删除_MEIxxxx文件夹。那么这句代码就把临时文件夹的地址赋值给base_path,再传入主函数中，那么运行文件时，程序就会在临时文件去寻找所需要的文件。
    else:
        base_path = os.path.abspath(".")#当前py文件的地址,当前目录的绝对路径
    return os.path.join(base_path, relative_path) #返回绝对地址

while 1:
    readpath = eg.fileopenbox('Where\'s your source?', 'PrintLabel 2.0v --Developer:SonicHuang')#用fileopenbox查找数据源
    if readpath == None: #如果点取消或者退出
        sys.exit()
    if os.path.splitext(readpath)[1] != '.xlsx': #根据扩展名判断选择的数据源是不是excel表格，openpyxl支持excel2003以后的版本，扩展名为xlsx
        eg.msgbox('CAUTION!:\nThe file is not Excel file or the Excel version is too old, the file extension must be xlsx.\nPlease select file again.', 'PrintLabel 2.0v --Developer:SonicHuang')#如果不是支持的版本，提示重新选择
        continue
    else: #如果扩展名正确则退出循环
        break
    
class Settings(EgStore): #这是easygui中记住用户设置的一个模块EgStore，用户设置之后通过store()这个方法在程序所在目录或者指定目录下生成一个txt的二进制文件，储存用的设置数据，通过restore()这个方法调取
    def __init__(self, filename): #参数是文件的绝对地址
        self.dimensions = ['28.3', '28.3', '28.3', '12.34', '12.34'] #初始化默认设置表格尺寸的数据
        self.fontstyle = ['宋体', '16', 'y', 'y','00BEBEBE', '宋体', '16', 'y', 'y', '00BEBEBE']#初始化默认设置字体的数据
        self.filename = filename
        self.restore() #调用储存的数据
settingsFilename = os.path.join(os.curdir, 'printlabelsettings.txt' )#设置储存数据的文本文件的绝对地址，不存在的话就新建
settings = Settings(settingsFilename) #实例化设置类并导入相关的默认的数据
dimensions = settings.dimensions #调用存储的数据
fontstyle = settings.fontstyle #调用储存的数据
msg = '<<...................Dimensions....................>>\nCheck the Excel Columnr/Row Setting Options for refrence.\nYou should input number from 8 to up for width, 14 to up for height'
title = 'PrintLabel 2.0v --Developer:SonicHuang'
fields = ['1st Row Height', '2nd Row Height', '3rd Row Height', 'A Column Width', 'B Column Width']
msg1 = '<<...................Font....................>>\nCheck the Excel Font Name and Size for your refrence\nColumn Color, check the Hex Color Code, example: Grey color is #BEBEBE, change # to 00, input 00BEBEBE'
fields1 = ['A Column Font Name', 'A Column Font Size','A Conlumn Bold(y/n)', 'A Conlumn Italic(y/n)', 'A Column Color', 'B Column Font Name', 'B Column Font Size', 'B Column Bold(y/n)', 'B Conlumn Italic(y/n)', 'B Column Color']

dimensions = eg.multenterbox(msg, title, fields, dimensions) #通过multenterbox设置数据
if dimensions == None:#如果被点了取消
    sys.exit()
fontstyle = eg.multenterbox(msg1, title, fields1, fontstyle)
if fontstyle == None: #如果被点了取消
    sys.exit()

settings.dimensions = dimensions #接受设置的数据
settings.fontstyle = fontstyle #接受设置的数据
settings.store() #储存设置的数据

   
dict1 = dict(zip(fields, dimensions)) #把项目和数据打包成字典，以方便调用
dict2 = dict(zip(fields1, fontstyle))
h1 = dict1['1st Row Height'] #调取数据
h2 = dict1['2nd Row Height']
h3 = dict1['3rd Row Height']
aw = dict1['A Column Width']
bw = dict1['B Column Width']
af = dict2['A Column Font Name']
bf = dict2['B Column Font Name']
afs = dict2['A Column Font Size']
bfs = dict2['B Column Font Size']
acc = dict2['A Column Color']
bcc = dict2['B Column Color']
acb = dict2['A Conlumn Bold(y/n)']#判断输入
if acb == 'y':
    acb = True
else:
    acb = None
bcb = dict2['B Column Bold(y/n)']
if bcb == 'y':
    bcb = True
else:
    bcb = None
aci = dict2['A Conlumn Italic(y/n)']
if aci == 'y':
    aci = True
else:
    aci = None
bci = dict2['B Conlumn Italic(y/n)']
if bci == 'y':
    bci = True
else:
    bci = None


wbr = opx.load_workbook(readpath) #调用openpyxl打开数据源返回的是一个类<openpyxl.workbook.workbook.Workbook object at 0x02AC7DF0>
sheetr = wbr.active #目前活动的sheet或者用wbr['sheetname']也可以.返回的是一个类<Worksheet "表1">，属于<class 'openpyxl.worksheet.worksheet.Worksheet'>
rownum = sheetr[1] #读取第一行的数据，返回的是一个单元格类组成的元组(<Cell '表1'.A1>, <Cell '表1'.B1>, <Cell '表1'.C1>, <Cell '表1'.D1>)


for r in range(2, sheetr.max_row + 1):#需要读取的数据行的范围，max_row是最大行数。这个循环是为了写入数据
    writepath = resource_path(os.path.join('printlabeltempfile', 'temp' + str(r-1) +'.xlsx'))#每读取一行的数据就会写入一个新建的excel临时表格。如果在同一个表格里面更替数据，并打印，最后传入打印机的是最后一组数据，所有打印出来的都是同一份数据。
    wbw = opx.Workbook()#定义一个工作薄。返回一个类<openpyxl.workbook.workbook.Workbook object at 0x0897FF70>
    sheetw = wbw.active #定义一个表单。返回一个类
    sheetw.row_dimensions[1].height = float(h1)#定义表单第一行的高度
    sheetw.row_dimensions[2].height = float(h2)#定义表单第二行的高度
    sheetw.row_dimensions[3].height = float(h3)#定义表单第三行的高度
    sheetw.column_dimensions['A'].width = aw #定义表单A列的宽度
    sheetw.column_dimensions['B'].width = bw #定义表单B列的宽度
    try:#这里用到接受异常的方法是为了让前面的循环被跳出后马上继续后面的一个循环，而不跳到主循环的开始位置
        i = 0
        for num in rownum: #迭代出第一行每个单元格的数据。这个循环是为了在A列先设置字体的大小颜色等，再设置对齐方式，水平垂直方向都居中，再依次传入数据
            i += 1 #rownum是第一行每个单元格类组成的元组,num是迭代出的单元格类储存有他的位置和值
            sheetw['A' + str(i)].font = Font(name = af, size = float(afs), color = acc, bold = acb, italic = aci)#定义字体
            sheetw['A' + str(i)].alignment = Alignment(horizontal = 'center', vertical = 'center')#定义居中
            sheetw['A' + str(i)] = num.value#写入单元格的值
            if i == 3:#只写入前三行的值
                raise Exception('nextround')#写完前三行立刻跳出循环，raise一个异常
    except Exception:#接受异常信息并进入下一面的程序，这里是循环，但不回到主循环
        
        n = 0 #这里是对B列的设置和赋值，
        for data in sheetr[r]:#迭出数据源每一行除第一行的所有单元格类的元组
            n += 1#先设置单元格字体大小颜色等，再设置居中方式
            sheetw['B' + str(n)].font = Font(name = bf, size = float(bfs), color = bcc, bold = bcb, italic = bci)
            sheetw['B' + str(n)].alignment = Alignment(horizontal = 'center', vertical = 'center')
            sheetw['B' + str(n)] = data.value #再写入对应的值
            if n == 3: #三行数据
                wbw.save(writepath)#写完数据保存表格
                os.startfile(writepath, 'print') #打印表格
                continue #进入主循环，再开始下一轮


askdel = eg.ccbox('<<...................CAUTION!....................>>\nPress any buttom after all the files\'r printed', 'PrintLabel 2.0v --Developer:SonicHuang', ('All files\'r printed', 'Exit this program'))
if askdel == True or askdel == False or askdel == None: #任务结束，系统自动删除临时文件
    pass #调用easygui来操作的目的是给打印一个时间，由于程序执行很快，如果还没有等数据出入打印机，临时表格就被删除，打印就会终止，程序会报错，所以这里用easygui来提示和暂时中止程序运行，等到打印完毕再执行删除任务。




